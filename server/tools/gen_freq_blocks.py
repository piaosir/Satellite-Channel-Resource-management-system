"""
生成 migration 011：
  - 从 Excel「通道规划状态」重新加载 frequency_block_realtime_status（规划分配块）
  - 从 Excel「通道占用状态」重新加载 occupation_realtime_status（频率分配块）
  - 修复 contract_record 的频率块外键引用

Excel 列位置（0-indexed）：
  「通道规划状态」（sheet index 8）
    0  规划记录ID         1  频率块标识-短          2  频率块（全码）
    3  维护备注-履约状态  4  维护备注-用户           5  维护备注-销售
    6  开关代码           7  矩阵状态号              8  偏移量
    9  占用宽度           10 状态(规划P/分配R/无效N) 11 状态修改时间
    12 用途               13 上行起频率              14 上行终止频率
    15 下行起频率         16 下行终止频率            17 所属卫星
    18 上行输入端口通道名 19 下行输出端口通道名
    20 规划块带宽(取)     21 规划块上行起频          22 规划块上行止频
    23 规划块下行起频     24 规划块下行止频          25 记录是否有效
    26 创建时间

  「通道占用状态」（sheet index 9）
    0  占用记录ID         1  频率块（全码/占用码）
    2  维护备注-履约状态  3  维护备注-用户           4  维护备注-销售
    5  开关代码           6  矩阵状态号              7  偏移量
    8  占用宽度           9  状态(规划P/分配R/无效N) 10 状态修改时间
    11 用途               12 上行起频率              13 上行终止频率
    14 下行起频率         15 下行终止频率            16 所属卫星
    17 上行输入端口通道名 18 下行输出端口通道名
    19 规划块带宽         20 规划块上行起频          21 规划块上行止频
    22 规划块下行起频     23 规划块下行止频          24 分配是否有效
    25 创建时间
"""
import argparse
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
MIGRATIONS = os.path.join(HERE, "..", "migrations")
DEFAULT_XLSX = os.path.join(HERE, "channel_data.xlsx")
BATCH = 200


def _blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() in ("", "#N/A", "#n/a"))


def txt(v) -> str:
    if _blank(v):
        return "NULL"
    s = str(v).strip()
    s = s.replace(";", "；").replace("'", "''")
    return "'" + s + "'"


def num(v) -> str:
    if _blank(v):
        return "NULL"
    if isinstance(v, bool):
        return "NULL"
    if isinstance(v, float) and v != v:
        return "NULL"
    if isinstance(v, (int, float)):
        return repr(float(v))
    s = str(v).strip().replace(",", "")
    try:
        return repr(float(s))
    except ValueError:
        return "NULL"


def is_invalid(v) -> bool:
    """返回 True 表示该行应跳过（无效）"""
    if v is None:
        return False
    s = str(v).strip()
    return s in ("否", "No", "0", "False", "N")


def emit_inserts(table: str, cols: list, rows_sql: list) -> str:
    col_list = "(" + ", ".join(f"`{c}`" for c in cols) + ")"
    out = []
    for i in range(0, len(rows_sql), BATCH):
        chunk = rows_sql[i:i + BATCH]
        out.append(
            f"INSERT INTO `{table}` {col_list} VALUES\n  "
            + ",\n  ".join(chunk) + ";"
        )
    return "\n".join(out)


def gen_planning_blocks(ws):
    """从「通道规划状态」表生成 frequency_block_realtime_status 的 INSERT SQL。"""
    cols = [
        "frequencyBlockCode", "frequencyBlockCode2", "switchCode",
        "frequencyOffset", "occupiedBandwidth", "partitionStatus",
        "usageType",
        "uplinkStartFreq", "uplinkEndFreq", "downlinkStartFreq", "downlinkEndFreq",
        "remarkFulfillment", "remarkUser", "remarkSales",
    ]
    rows_sql = []
    skipped = 0
    # build lookup: switchCode → frequencyBlockCode2 (for occupation FK later)
    switch_to_plan_code = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        validity = row[25]   # 记录是否有效
        status   = row[10]   # 状态
        if is_invalid(validity):
            skipped += 1
            continue
        if not _blank(status) and str(status).strip() == "N":
            skipped += 1
            continue
        sw_code  = row[6]    # 开关代码
        fb_short = row[1]    # 频率块标识-短
        fb_full  = row[2]    # 频率块（全码）
        if _blank(sw_code) and _blank(fb_short) and _blank(fb_full):
            skipped += 1
            continue
        # 记录 switch → 规划块全码（供占用记录 FK 回填使用）
        # key 用三元组 (switchCode, ul_start, ul_end) 避免同一开关多个规划块碰撞
        if not _blank(sw_code) and not _blank(fb_full):
            ul_start = row[13]
            ul_end   = row[14]
            try:
                ul_s = float(ul_start) if ul_start is not None else None
                ul_e = float(ul_end)   if ul_end   is not None else None
            except (ValueError, TypeError):
                ul_s, ul_e = None, None
            key = (str(sw_code).strip(), ul_s, ul_e)
            switch_to_plan_code[key] = str(fb_full).strip() if fb_full else None

        rows_sql.append("(" + ", ".join([
            txt(fb_short),
            txt(fb_full),
            txt(sw_code),
            num(row[8]),    # 偏移量
            num(row[9]),    # 占用宽度
            txt(status),    # 状态
            txt(row[12]),   # 用途
            num(row[13]),   # 上行起频率
            num(row[14]),   # 上行终止频率
            num(row[15]),   # 下行起频率
            num(row[16]),   # 下行终止频率
            txt(row[3]),    # 维护备注-履约状态
            txt(row[4]),    # 维护备注-用户
            txt(row[5]),    # 维护备注-销售
        ]) + ")")

    print(f"  规划块有效行: {len(rows_sql)}, 跳过: {skipped}")
    return rows_sql, cols, switch_to_plan_code


def gen_occupation_records(ws, switch_to_plan_code: dict):
    """从「通道占用状态」表生成 occupation_realtime_status 的 INSERT SQL。"""
    cols = [
        "occupationCode", "switchCode",
        "frequencyOffset", "occupiedBandwidth", "partitionStatus",
        "usageType",
        "uplinkStartFreq", "uplinkEndFreq", "downlinkStartFreq", "downlinkEndFreq",
        "remarkFulfillment", "remarkUser", "remarkSales",
        "planningBlockCode",   # 冗余存储，用于后续 JOIN 更新 planningBlockId
        "isValid",
    ]
    rows_sql = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        validity = row[24]   # 分配是否有效
        status   = row[9]    # 状态
        if is_invalid(validity):
            skipped += 1
            continue
        if not _blank(status) and str(status).strip() == "N":
            skipped += 1
            continue
        occ_code = row[1]    # 频率块（全码/占用码）
        sw_code  = row[5]    # 开关代码
        if _blank(occ_code) and _blank(sw_code):
            skipped += 1
            continue

        # 查找所属规划块代码：通过开关代码 + 入通道起/止频率（对应规划块上/止行频率）
        plan_code_val = None
        if not _blank(sw_code):
            plan_ul_start = row[20]  # 索引"入通道起频率" = 规划块上行起频
            plan_ul_end   = row[21]  # 索引"入通道止频率" = 规划块上行止频
            try:
                ul_s = float(plan_ul_start) if not _blank(plan_ul_start) else None
                ul_e = float(plan_ul_end)   if not _blank(plan_ul_end)   else None
            except (ValueError, TypeError):
                ul_s, ul_e = None, None
            key = (str(sw_code).strip(), ul_s, ul_e)
            plan_code_val = switch_to_plan_code.get(key)
            if plan_code_val is None:
                # 回退：只用起频（兼容止频缺失的情况）
                for k, v in switch_to_plan_code.items():
                    if k[0] == str(sw_code).strip() and k[1] == ul_s:
                        plan_code_val = v
                        break
            if plan_code_val is None:
                # 最终回退：仅用开关代码
                for k, v in switch_to_plan_code.items():
                    if k[0] == str(sw_code).strip():
                        plan_code_val = v
                        break

        rows_sql.append("(" + ", ".join([
            txt(occ_code),   # occupationCode
            txt(sw_code),    # switchCode
            num(row[7]),     # 偏移量
            num(row[8]),     # 占用宽度
            txt(status),     # 状态
            txt(row[11]),    # 用途
            num(row[12]),    # 上行起频率
            num(row[13]),    # 上行终止频率
            num(row[14]),    # 下行起频率
            num(row[15]),    # 下行终止频率
            txt(row[2]),     # 维护备注-履约状态
            txt(row[3]),     # 维护备注-用户
            txt(row[4]),     # 维护备注-销售
            txt(plan_code_val),  # planningBlockCode
            "'1'",           # isValid
        ]) + ")")

    print(f"  占用记录有效行: {len(rows_sql)}, 跳过: {skipped}")
    return rows_sql, cols


def gen_migration(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    names = wb.sheetnames

    # 找到两个目标 Sheet（按 index 8/9，也支持按名称搜索）
    ws_plan = None
    ws_occ  = None
    for name in names:
        if "通道规划状态" in name:
            ws_plan = wb[name]
        if "通道分配状态" in name or "通道占用状态" in name:
            ws_occ = wb[name]
    # 找不到时按索引兜底
    if ws_plan is None and len(names) > 8:
        ws_plan = wb[names[8]]
        print(f"  回退：使用 sheet[8] = {names[8]}")
    if ws_occ is None and len(names) > 9:
        ws_occ  = wb[names[9]]
        print(f"  回退：使用 sheet[9] = {names[9]}")

    if ws_plan is None:
        raise ValueError("未找到「通道规划状态」sheet")
    if ws_occ is None:
        raise ValueError("未找到「通道占用状态」sheet")

    print("生成规划块数据…")
    plan_rows, plan_cols, switch_to_plan_code = gen_planning_blocks(ws_plan)
    print("生成占用记录数据…")
    occ_rows, occ_cols = gen_occupation_records(ws_occ, switch_to_plan_code)

    wb.close()

    body_lines = [
        "-- 011: 修正规划分配块与频率分配块数据",
        "--   frequency_block_realtime_status ← 通道规划状态（规划分配块）",
        "--   occupation_realtime_status       ← 通道占用状态（频率分配块）",
        "",
        "-- ── 重置规划分配块（frequency_block_realtime_status）──",
        "-- 清空关联占用记录的外键（先置 NULL，再重建）",
        "UPDATE `occupation_realtime_status` SET `planningBlockId` = NULL, `planningBlockCode` = NULL;",
        "-- 清空规划块数据",
        "DELETE FROM `frequency_block_realtime_status`;",
        "",
        emit_inserts("frequency_block_realtime_status", plan_cols, plan_rows),
        "",
        "-- 解析 switchId",
        "UPDATE `frequency_block_realtime_status` fb",
        "JOIN `matrix_switch_status` sw ON sw.`switchCode` = fb.`switchCode`",
        "SET fb.`switchId` = sw.`id`;",
        "",
        "-- ── 重置频率分配块（occupation_realtime_status）──",
        "DELETE FROM `occupation_realtime_status`;",
        "",
        emit_inserts("occupation_realtime_status", occ_cols, occ_rows),
        "",
        "-- 解析 switchId",
        "UPDATE `occupation_realtime_status` o",
        "JOIN `matrix_switch_status` sw ON sw.`switchCode` = o.`switchCode`",
        "SET o.`switchId` = sw.`id`;",
        "",
        "-- 回填 planningBlockId（按 planningBlockCode → frequency_block_realtime_status.frequencyBlockCode2）",
        "UPDATE `occupation_realtime_status` o",
        "JOIN `frequency_block_realtime_status` f ON f.`frequencyBlockCode2` = o.`planningBlockCode`",
        "SET o.`planningBlockId` = f.`id`;",
        "",
        "-- ── 修复 contract_record 频率块外键引用（指向分配块）──",
        "UPDATE `contract_record` c",
        "JOIN `occupation_realtime_status` o ON o.`occupationCode` = c.`frequencyBlockCode2`",
        "SET c.`frequencyBlockId` = o.`id`;",
        "",
        "-- 同时尝试用规划块代码匹配（兜底）",
        "UPDATE `contract_record` c",
        "JOIN `frequency_block_realtime_status` f ON f.`frequencyBlockCode2` = c.`frequencyBlockCode2`",
        "SET c.`frequencyBlockId` = f.`id`",
        "WHERE c.`frequencyBlockId` IS NULL;",
    ]

    body = "\n".join(body_lines)
    out_path = os.path.join(MIGRATIONS, "011_fix_planning_and_occupation_data.sql")
    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(body.rstrip() + "\n")
    print(f"写出 {out_path}  ({len(body.splitlines())} 行)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=DEFAULT_XLSX)
    args = ap.parse_args()
    print(f"读取 Excel：{args.excel}")
    gen_migration(args.excel)
    print("完成。")


if __name__ == "__main__":
    main()
