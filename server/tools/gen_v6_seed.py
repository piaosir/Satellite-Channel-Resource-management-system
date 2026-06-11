# -*- coding: utf-8 -*-
"""
v6 建库 + 灌数生成器
====================
读取领导定稿的《数据库建表数据-完整版-20260604(清理空行)》Excel,
生成 server/migrations/ 下的全部迁移文件:

    001_create_schema_v6.sql      19 张表 DDL
    002_seed_resource_layer.sql   资源层:卫星/信标/通道组/通道/矩阵/端口/开关
    003_seed_status_layer.sql     状态层:通道规划状态/通道分配状态
    004_seed_business_layer.sql   业务层:客户/用户/合约/交付记录/自有三表

设计要点(对应《数据库整体逻辑说明-给开发同事.txt》):
  - 所有主键沿用 Excel 中的原始 ID,关联解析(satelliteId/channelId/
    planningBlockId/allocationId 等)在本脚本内存中完成后直接写入
  - 块代码一律 trim 首尾空格;同时解析出卫星/带宽/上下行极化/波束/
    起止频率等结构化列,便于查询
  - "占用/释放"与"分配是否有效"互相独立,不做任何联动
  - 空表(三张切换日志/自有业务系统/载波清单)只建结构不灌数

用法:
    cd server/tools
    python gen_v6_seed.py
"""

import os
import re
import sys
from datetime import datetime, date

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "db_data_20260604.xlsx")
MIG_DIR = os.path.join(HERE, "..", "migrations")

BATCH = 200  # 每条 INSERT 语句的行数

# ── 块代码解析 ─────────────────────────────────────────────────
# 格式: 卫星_BW带宽_U极化+波束(5位定宽,左侧下划线补齐)_S起_E止_D..._S起_E止
BLOCK_RE = re.compile(
    r"^(\w+?)_BW([\d.]+)"
    r"_U(.)(.{5})_S([\d.]+)_E([\d.]+)"
    r"_D(.)(.{5})_S([\d.]+)_E([\d.]+)$"
)


def parse_block(code):
    """解析块代码,返回 dict;解析失败返回 None。"""
    m = BLOCK_RE.match(code)
    if not m:
        return None
    sat, bw, up, ub, us, ue, dp, db, ds, de = m.groups()
    return {
        "sat": sat, "bw": float(bw),
        "upol": up, "ubeam": ub.strip("_"), "ustart": float(us), "uend": float(ue),
        "dpol": dp, "dbeam": db.strip("_"), "dstart": float(ds), "dend": float(de),
    }


# ── SQL 字面量 ─────────────────────────────────────────────────
def sq(v):
    """字符串 → SQL 字面量(trim;转义引号/反斜杠;ASCII 分号替换为全角,
    保护 migrate.py 的语句切分)。"""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return "NULL"
    s = s.replace("\\", "\\\\").replace("'", "''")
    s = s.replace("\r\n", "\\n").replace("\n", "\\n").replace("\r", "\\n")
    return f"'{s}'"


def num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return "NULL"
    try:
        f = float(str(v).strip())
    except ValueError:
        return "NULL"
    return str(int(f)) if f == int(f) else repr(f)


def intv(v):
    n = num(v)
    if n == "NULL":
        return n
    return str(int(float(n)))


def dt(v):
    """datetime → 'YYYY-MM-DD HH:MM:SS';date → 'YYYY-MM-DD'。"""
    if v is None or (isinstance(v, str) and not v.strip()):
        return "NULL"
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(v, date):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return sq(v)


def d_only(v):
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    return sq(v)


def yn(v):
    """是/否 → 1/0。"""
    if v is None:
        return "NULL"
    s = str(v).strip()
    return "1" if s == "是" else ("0" if s == "否" else "NULL")


# ── Excel 读取 ─────────────────────────────────────────────────
def rows_of(wb, sheet):
    """返回 [dict] 形式的全部数据行(以首行表头为键,值已去 str 两端空格)。"""
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if h is None:
                continue
            d[h] = v.strip() if isinstance(v, str) else v
        out.append(d)
    return out


def uniquify_ids(rows, idcol, label):
    """Excel 手填序号存在重号(不同段落编号重叠),后出现的重号行顺延为
    max+1 重新编号。其余表均按代号关联,重编号不影响引用。"""
    seen, nxt, n_fix = set(), 0, 0
    for r in rows:
        if r[idcol] is not None:
            nxt = max(nxt, int(r[idcol]))
    for r in rows:
        v = int(r[idcol]) if r[idcol] is not None else None
        if v is None or v in seen:
            nxt += 1
            r[idcol] = nxt
            n_fix += 1
        else:
            seen.add(v)
    if n_fix:
        print(f"  [修复] {label}: {n_fix} 行 {idcol} 重号/缺失,已顺延重编号")
    return rows


def emit(table, cols, value_rows):
    """生成分批 INSERT 语句。"""
    if not value_rows:
        return f"-- {table}: 无种子数据(空表,结构先行)\n"
    col_list = "(" + ", ".join(f"`{c}`" for c in cols) + ")"
    parts = []
    for i in range(0, len(value_rows), BATCH):
        chunk = value_rows[i:i + BATCH]
        parts.append(
            f"INSERT INTO `{table}` {col_list} VALUES\n  " + ",\n  ".join(chunk) + ";"
        )
    return "\n".join(parts) + "\n"


def write(name, body):
    os.makedirs(MIG_DIR, exist_ok=True)
    path = os.path.join(MIG_DIR, name)
    with open(path, "w", encoding="utf-8") as f:
        f.write(body)
    print(f"  写出 {name}  ({len(body) // 1024} KB)")


# ══════════════════════════════════════════════════════════════
#  001  Schema DDL
# ══════════════════════════════════════════════════════════════
SCHEMA = """-- 001: v6 全新建表(19 张,对应《数据库建表数据-完整版-20260604》)
-- 命名沿用项目惯例:冗余 code 字段 + 解析 id 字段(后缀注释"解析")

-- ── 资源层 ────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS `satellite_info` (
  `id` INT NOT NULL COMMENT '卫星清单ID(沿用Excel)',
  `satelliteCode` VARCHAR(64) NOT NULL COMMENT '卫星代号',
  `satelliteCodeNonStd` VARCHAR(64) NULL COMMENT '卫星代号(非标)',
  `satelliteName` VARCHAR(128) NULL COMMENT '卫星名称',
  `statusText` VARCHAR(32) NULL COMMENT '状态:在轨运营/在轨停服/离轨',
  `orbitPosition` VARCHAR(64) NULL COMMENT '轨位',
  `launchDate` DATE NULL COMMENT '发射时间',
  `designLife` VARCHAR(32) NULL COMMENT '寿命(年)',
  `manufacturer` VARCHAR(128) NULL COMMENT '制造商',
  `platform` VARCHAR(128) NULL COMMENT '卫星平台',
  `coverage` TEXT NULL COMMENT '覆盖',
  `payload` TEXT NULL COMMENT '有效载荷',
  PRIMARY KEY (`id`),
  UNIQUE KEY `uk_satellite_code` (`satelliteCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='卫星清单';

CREATE TABLE IF NOT EXISTS `beacon_info` (
  `id` INT NOT NULL COMMENT '信标ID(沿用Excel)',
  `satelliteCode` VARCHAR(64) NULL COMMENT '关联卫星代号',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `band` VARCHAR(32) NULL COMMENT '频段:C/Ku/EKu/Ka',
  `polarization` VARCHAR(16) NULL COMMENT '极化:H/V/L/R',
  `frequency` DECIMAL(12,2) NULL COMMENT '频点(MHz)',
  PRIMARY KEY (`id`),
  KEY `idx_beacon_satellite` (`satelliteId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='信标清单';

CREATE TABLE IF NOT EXISTS `channel_group_info` (
  `id` INT NOT NULL COMMENT '通道组ID(沿用Excel)',
  `channelGroupCode` VARCHAR(128) NOT NULL COMMENT '通道组代号',
  `groupSeq` VARCHAR(32) NULL COMMENT '通道组序号',
  `satelliteCode` VARCHAR(64) NULL COMMENT '关联卫星代号',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `antennaName` VARCHAR(128) NULL COMMENT '波束(天线)名称',
  `antennaCode` VARCHAR(32) NULL COMMENT '波束(天线)代号',
  `txRxType` VARCHAR(8) NULL COMMENT '收/发:R/T',
  `polarization` VARCHAR(16) NULL COMMENT '极化:H/V/L/R/X/Y',
  `band` VARCHAR(32) NULL COMMENT '频率(频段):C/EC/Ku/EKu/KuBSS/规划Ku/Ka/KaBSS',
  `channelCount` INT NULL COMMENT '通道数',
  `primaryReceiverCode` VARCHAR(64) NULL COMMENT '主份接收机代码',
  `backupReceiverCode1` VARCHAR(64) NULL COMMENT '一备接收机代码',
  `backupReceiverCode2` VARCHAR(64) NULL COMMENT '二备接收机代码',
  `receiverActiveStatus` VARCHAR(16) NULL COMMENT '接收机主备状态',
  PRIMARY KEY (`id`),
  UNIQUE KEY `uk_channel_group_code` (`channelGroupCode`),
  KEY `idx_group_satellite` (`satelliteId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='通道组清单(含接收机主备)';

CREATE TABLE IF NOT EXISTS `channel_info` (
  `id` INT NOT NULL COMMENT '通道ID(沿用Excel)',
  `channelCode` VARCHAR(128) NOT NULL COMMENT '通道代号(唯一键)',
  `channelFullName` VARCHAR(128) NULL COMMENT '通道全称',
  `channelShortName` VARCHAR(64) NULL COMMENT '通道简称(非唯一)',
  `commonName` VARCHAR(128) NULL COMMENT '常用名',
  `channelGroupCode` VARCHAR(128) NULL COMMENT '关联通道组代号',
  `channelGroupId` INT NULL COMMENT '通道组id(解析)',
  `channelBandwidth` DECIMAL(12,2) NULL COMMENT '通道带宽(MHz)',
  `channelStartFreq` DECIMAL(12,2) NULL COMMENT '通道起始频率(MHz)',
  `channelEndFreq` DECIMAL(12,2) NULL COMMENT '通道终止频率(MHz)',
  PRIMARY KEY (`id`),
  UNIQUE KEY `uk_channel_code` (`channelCode`),
  KEY `idx_channel_group` (`channelGroupId`),
  KEY `idx_channel_short_name` (`channelShortName`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='通道清单';

CREATE TABLE IF NOT EXISTS `switch_matrix_info` (
  `id` INT NOT NULL COMMENT '矩阵ID(沿用Excel)',
  `matrixCode` VARCHAR(128) NOT NULL COMMENT '矩阵代码',
  `satelliteCode` VARCHAR(64) NULL COMMENT '关联卫星代号',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `matrixType` TINYINT NULL COMMENT '类型:1常规开关矩阵 2 DTP大矩阵',
  `matrixSeq` INT NULL COMMENT '序号',
  `inputPortCount` INT NULL COMMENT '输入端口数',
  `outputPortCount` INT NULL COMMENT '输出端口数',
  `effectiveStatus` TINYINT NULL COMMENT '生效状态:1有效 0无效',
  `remark` VARCHAR(500) NULL COMMENT '备注',
  `updateTime` DATETIME NULL COMMENT '变更时间',
  PRIMARY KEY (`id`),
  UNIQUE KEY `uk_matrix_code` (`matrixCode`),
  KEY `idx_matrix_satellite` (`satelliteId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='矩阵清单';

CREATE TABLE IF NOT EXISTS `matrix_port_info` (
  `id` INT NOT NULL COMMENT '端口ID(沿用Excel)',
  `portCode` VARCHAR(128) NULL COMMENT '端口代码',
  `matrixCode` VARCHAR(128) NULL COMMENT '关联矩阵代码',
  `matrixId` INT NULL COMMENT '矩阵id(解析)',
  `ioType` VARCHAR(8) NULL COMMENT '输入/输出:I/O',
  `portSeq` INT NULL COMMENT '序号',
  `channelShortName` VARCHAR(64) NULL COMMENT '关联通道代码(=通道简称)',
  `channelId` INT NULL COMMENT '通道id(解析;同卫星内简称多义时为NULL)',
  PRIMARY KEY (`id`),
  KEY `idx_port_matrix` (`matrixId`),
  KEY `idx_port_channel` (`channelId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='矩阵端口清单';

CREATE TABLE IF NOT EXISTS `matrix_switch_status` (
  `id` INT NOT NULL COMMENT '开关ID(沿用Excel)',
  `switchCode` VARCHAR(128) NOT NULL COMMENT '开关代码',
  `matrixCode` VARCHAR(128) NULL COMMENT '关联矩阵代码',
  `matrixId` INT NULL COMMENT '矩阵id(解析)',
  `inputPortSeq` INT NULL COMMENT '入端口号',
  `outputPortSeq` INT NULL COMMENT '出端口号',
  `inputPortId` INT NULL COMMENT '入端口id(解析)',
  `outputPortId` INT NULL COMMENT '出端口id(解析)',
  `switchType` VARCHAR(16) NULL COMMENT '开关是否可切:常通/可切',
  `switchStatus` TINYINT NULL COMMENT '开关状态:1通 0断',
  `primaryAmpCode` VARCHAR(64) NULL COMMENT '主份放大器代码',
  `backupAmpCode1` VARCHAR(64) NULL COMMENT '一备放大器代码',
  `backupAmpCode2` VARCHAR(64) NULL COMMENT '二备放大器代码',
  `ampActiveStatus` VARCHAR(16) NULL COMMENT '放大器主备状态:P0主份/P1一备/P2二备',
  `updateTime` DATETIME NULL COMMENT '变更时间',
  PRIMARY KEY (`id`),
  UNIQUE KEY `uk_switch_code` (`switchCode`),
  KEY `idx_switch_matrix` (`matrixId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='矩阵开关状态(实时,含放大器主备)';

-- ── 切换日志(三张同构,结构先行) ────────────────────────────
CREATE TABLE IF NOT EXISTS `matrix_switch_log` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '切换日志ID',
  `switchCode` VARCHAR(128) NULL COMMENT '开关代码',
  `switchId` INT NULL COMMENT '开关id(解析)',
  `beforeStatus` VARCHAR(32) NULL COMMENT '切换前状态',
  `afterStatus` VARCHAR(32) NULL COMMENT '切换后状态',
  `switchTime` DATETIME NULL COMMENT '切换时间',
  `operator` VARCHAR(64) NULL COMMENT '操作人',
  `registrar` VARCHAR(64) NULL COMMENT '登记人',
  PRIMARY KEY (`id`),
  KEY `idx_mslog_switch` (`switchCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='矩阵开关切换日志';

CREATE TABLE IF NOT EXISTS `amplifier_switch_log` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '切换日志ID',
  `switchCode` VARCHAR(128) NULL COMMENT '开关代码',
  `switchId` INT NULL COMMENT '开关id(解析)',
  `beforeStatus` VARCHAR(32) NULL COMMENT '切换前状态',
  `afterStatus` VARCHAR(32) NULL COMMENT '切换后状态',
  `switchTime` DATETIME NULL COMMENT '切换时间',
  `operator` VARCHAR(64) NULL COMMENT '操作人',
  `registrar` VARCHAR(64) NULL COMMENT '登记人',
  PRIMARY KEY (`id`),
  KEY `idx_amplog_switch` (`switchCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='放大器切换日志';

CREATE TABLE IF NOT EXISTS `receiver_switch_log` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '切换日志ID',
  `channelGroupCode` VARCHAR(128) NULL COMMENT '通道组代号',
  `channelGroupId` INT NULL COMMENT '通道组id(解析)',
  `beforeStatus` VARCHAR(32) NULL COMMENT '切换前状态',
  `afterStatus` VARCHAR(32) NULL COMMENT '切换后状态',
  `switchTime` DATETIME NULL COMMENT '切换时间',
  `operator` VARCHAR(64) NULL COMMENT '操作人',
  `registrar` VARCHAR(64) NULL COMMENT '登记人',
  PRIMARY KEY (`id`),
  KEY `idx_rcvlog_group` (`channelGroupCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='接收机_变频器切换日志';

-- ── 状态层 ────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS `channel_planning_status` (
  `id` INT NOT NULL COMMENT '规划记录ID(沿用Excel)',
  `blockCode` VARCHAR(128) NOT NULL COMMENT '块代码(已trim)',
  `usageType` VARCHAR(16) NULL COMMENT '用途:自用/出租/合作/禁用',
  `isValid` TINYINT NULL COMMENT '记录是否有效:1是 0否',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  `satelliteCode` VARCHAR(64) NULL COMMENT '卫星代号(解析自块代码)',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `bandwidth` DECIMAL(12,2) NULL COMMENT '带宽MHz(解析)',
  `uplinkPolarization` VARCHAR(8) NULL COMMENT '上行极化(解析)',
  `uplinkBeam` VARCHAR(16) NULL COMMENT '上行波束(解析)',
  `uplinkStartFreq` DECIMAL(12,2) NULL COMMENT '上行起始频率(解析)',
  `uplinkEndFreq` DECIMAL(12,2) NULL COMMENT '上行终止频率(解析)',
  `downlinkPolarization` VARCHAR(8) NULL COMMENT '下行极化(解析)',
  `downlinkBeam` VARCHAR(16) NULL COMMENT '下行波束(解析)',
  `downlinkStartFreq` DECIMAL(12,2) NULL COMMENT '下行起始频率(解析)',
  `downlinkEndFreq` DECIMAL(12,2) NULL COMMENT '下行终止频率(解析)',
  `channelId` INT NULL COMMENT '所落接收通道id(解析:上行频率范围归属)',
  PRIMARY KEY (`id`),
  KEY `idx_plan_block_code` (`blockCode`),
  KEY `idx_plan_satellite` (`satelliteId`),
  KEY `idx_plan_channel` (`channelId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='通道规划状态(基底:块+用途)';

CREATE TABLE IF NOT EXISTS `channel_allocation_status` (
  `id` INT NOT NULL COMMENT '占用记录ID(沿用Excel)',
  `blockCode` VARCHAR(128) NOT NULL COMMENT '块代码(已trim)',
  `isValid` TINYINT NULL COMMENT '分配是否有效:1是 0否(仅拆分/冲突时置0,与占用释放无关)',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  `satelliteCode` VARCHAR(64) NULL COMMENT '卫星代号(解析自块代码)',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `bandwidth` DECIMAL(12,2) NULL COMMENT '带宽MHz(解析)',
  `uplinkPolarization` VARCHAR(8) NULL COMMENT '上行极化(解析)',
  `uplinkBeam` VARCHAR(16) NULL COMMENT '上行波束(解析)',
  `uplinkStartFreq` DECIMAL(12,2) NULL COMMENT '上行起始频率(解析)',
  `uplinkEndFreq` DECIMAL(12,2) NULL COMMENT '上行终止频率(解析)',
  `downlinkPolarization` VARCHAR(8) NULL COMMENT '下行极化(解析)',
  `downlinkBeam` VARCHAR(16) NULL COMMENT '下行波束(解析)',
  `downlinkStartFreq` DECIMAL(12,2) NULL COMMENT '下行起始频率(解析)',
  `downlinkEndFreq` DECIMAL(12,2) NULL COMMENT '下行终止频率(解析)',
  `planningBlockId` INT NULL COMMENT '所属规划块id(解析:同星同极化同波束范围包含)',
  `channelId` INT NULL COMMENT '所落接收通道id(解析)',
  PRIMARY KEY (`id`),
  KEY `idx_alloc_block_code` (`blockCode`),
  KEY `idx_alloc_satellite` (`satelliteId`),
  KEY `idx_alloc_planning` (`planningBlockId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='通道分配状态(实际占用快照)';

-- ── 业务层 ────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS `customer_info` (
  `customerCode` VARCHAR(32) NOT NULL COMMENT '客户ID(KH-xxxxxxx)',
  `customerName` VARCHAR(255) NULL COMMENT '客户全称(存在重名,以customerCode为准)',
  `creditCode` VARCHAR(64) NULL COMMENT '统一社会信用代码',
  `status` TINYINT NULL COMMENT '客户状态',
  `createdTime` DATETIME NULL COMMENT '建档时间',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  PRIMARY KEY (`customerCode`),
  KEY `idx_customer_name` (`customerName`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='客户信息';

CREATE TABLE IF NOT EXISTS `user_info` (
  `id` INT NOT NULL COMMENT '用户ID(沿用Excel,与合约一一对应)',
  `customerCode` VARCHAR(32) NULL COMMENT '所属客户ID',
  `customerName` VARCHAR(255) NULL COMMENT '所属客户全称(冗余)',
  `status` TINYINT NULL COMMENT '用户状态',
  `createdTime` DATETIME NULL COMMENT '建档时间',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  PRIMARY KEY (`id`),
  KEY `idx_user_customer` (`customerCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='用户信息(客户与合约的中间实体)';

CREATE TABLE IF NOT EXISTS `bandwidth_contract_info` (
  `id` INT NOT NULL COMMENT '合约ID(沿用Excel)',
  `customerName` VARCHAR(255) NULL COMMENT '所属客户(名称)',
  `customerCode` VARCHAR(32) NULL COMMENT '客户ID(解析:经用户表)',
  `userId` INT NULL COMMENT '用户号(=user_info.id,一一对应)',
  `mainOrderCode` VARCHAR(64) NULL COMMENT '所属主订单',
  `productName` VARCHAR(128) NULL COMMENT '签约带宽商品',
  `productType` VARCHAR(32) NULL COMMENT '商品类型:长租等',
  `bandwidthMHz` DECIMAL(12,2) NULL COMMENT '签约带宽权益(MHz)',
  `divisibleBlockCount` INT NULL COMMENT '可分频率块数量',
  `periods` DECIMAL(8,2) NULL COMMENT '期数',
  `amount` DECIMAL(14,2) NULL COMMENT '签约金额',
  `startTime` DATETIME NULL COMMENT '开通时间',
  `endTime` DATETIME NULL COMMENT '到期时间',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  PRIMARY KEY (`id`),
  KEY `idx_contract_user` (`userId`),
  KEY `idx_contract_customer` (`customerCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='(带宽)合约清单';

CREATE TABLE IF NOT EXISTS `contract_delivery_record` (
  `id` INT NOT NULL COMMENT '交付记录ID(沿用Excel)',
  `contractId` INT NULL COMMENT '关联合约ID',
  `blockCode` VARCHAR(128) NULL COMMENT '频率块代码(必须引用通道分配状态,已trim)',
  `allocationId` INT NULL COMMENT '分配块id(解析:块代码精确匹配)',
  `exclusiveType` VARCHAR(16) NULL COMMENT '独占/共享',
  `satelliteCode` VARCHAR(64) NULL COMMENT '频率块所属卫星',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `bandwidth` DECIMAL(12,2) NULL COMMENT '交付频率块带宽(MHz)',
  `action` VARCHAR(16) NULL COMMENT '动作:占用/释放',
  `actionTime` DATETIME NULL COMMENT '时间',
  `handler` VARCHAR(64) NULL COMMENT '受理人员',
  `registrar` VARCHAR(64) NULL COMMENT '登记人员',
  PRIMARY KEY (`id`),
  KEY `idx_delivery_contract` (`contractId`),
  KEY `idx_delivery_block` (`blockCode`),
  KEY `idx_delivery_allocation` (`allocationId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='(带宽)合约-交付过程记录';

CREATE TABLE IF NOT EXISTS `own_business_system_info` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '业务系统ID',
  `systemCode` VARCHAR(128) NULL COMMENT '业务系统代称',
  `basebandName` VARCHAR(128) NULL COMMENT '基带系统名称',
  `createdTime` DATETIME NULL COMMENT '建档时间',
  `updateTime` DATETIME NULL COMMENT '最后更新时间',
  PRIMARY KEY (`id`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='自有业务系统信息(结构先行)';

CREATE TABLE IF NOT EXISTS `own_carrier_info` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '载波ID',
  `businessSystemId` INT NULL COMMENT '所属业务系统ID',
  `direction` VARCHAR(16) NULL COMMENT '前向/返向',
  `bandwidth` DECIMAL(12,2) NULL COMMENT '载波带宽(MHz)',
  PRIMARY KEY (`id`),
  KEY `idx_carrier_system` (`businessSystemId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='(自有)载波清单(结构先行)';

CREATE TABLE IF NOT EXISTS `own_carrier_usage_record` (
  `id` INT NOT NULL COMMENT '使用记录ID(沿用Excel)',
  `carrierId` INT NULL COMMENT '关联载波ID(当前数据为空)',
  `blockCode` VARCHAR(128) NULL COMMENT '频率块代码(必须引用通道分配状态,已trim)',
  `allocationId` INT NULL COMMENT '分配块id(解析:块代码精确匹配)',
  `exclusiveType` VARCHAR(16) NULL COMMENT '独占/共享',
  `satelliteCode` VARCHAR(64) NULL COMMENT '频率块所属卫星',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `bandwidth` DECIMAL(12,2) NULL COMMENT '频率块带宽(MHz)',
  `action` VARCHAR(16) NULL COMMENT '动作:占用/释放',
  `actionTime` DATETIME NULL COMMENT '时间',
  `handler` VARCHAR(64) NULL COMMENT '受理人员',
  `registrar` VARCHAR(64) NULL COMMENT '登记人员',
  PRIMARY KEY (`id`),
  KEY `idx_usage_carrier` (`carrierId`),
  KEY `idx_usage_block` (`blockCode`),
  KEY `idx_usage_allocation` (`allocationId`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='(自有)载波使用带宽过程记录';
"""


# ══════════════════════════════════════════════════════════════
def main():
    print(f"读取 {os.path.basename(XLSX)} ...")
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    # ── 资源层原始行 ──
    sats = rows_of(wb, "卫星清单")
    beacons = rows_of(wb, "信标清单")
    groups = rows_of(wb, "通道组清单")
    channels = rows_of(wb, "通道清单")
    matrices = rows_of(wb, "矩阵清单")
    ports = rows_of(wb, "矩阵端口清单")
    switches = rows_of(wb, "矩阵开关状态（实时）")
    plans = rows_of(wb, "通道规划状态")
    allocs = rows_of(wb, "通道分配状态")
    contracts = rows_of(wb, "（带宽）合约清单")
    deliveries = rows_of(wb, "（带宽）合约-交付过程记录")
    customers = rows_of(wb, "客户信息")
    users = rows_of(wb, "用户信息")
    usages = rows_of(wb, "（自有）载波使用带宽过程记录")

    # Excel 手填 ID 可能重号,统一检查修复(其余关联均按代号,安全)
    uniquify_ids(sats, "卫星清单ID", "卫星清单")
    uniquify_ids(beacons, "信标ID", "信标清单")
    uniquify_ids(groups, "通道组ID", "通道组清单")
    uniquify_ids(channels, "通道ID", "通道清单")
    uniquify_ids(matrices, "矩阵ID", "矩阵清单")
    uniquify_ids(ports, "端口ID", "矩阵端口清单")
    uniquify_ids(switches, "开关ID", "矩阵开关状态")
    uniquify_ids(plans, "规划记录ID", "通道规划状态")
    uniquify_ids(allocs, "占用记录ID", "通道分配状态")
    uniquify_ids(contracts, "合约ID", "合约清单")
    uniquify_ids(deliveries, "交付记录ID", "交付过程记录")
    uniquify_ids(users, "用户ID", "用户信息")
    uniquify_ids(usages, "使用记录ID", "载波使用记录")

    # ── 解析映射 ──
    sat_by_code = {str(s["卫星代号"]).strip(): int(s["卫星清单ID"]) for s in sats}
    group_by_code = {str(g["通道组代号"]).strip(): int(g["通道组ID"]) for g in groups}
    matrix_by_code = {str(m["矩阵代码"]).strip(): int(m["矩阵ID"]) for m in matrices}
    matrix_sat = {int(m["矩阵ID"]): str(m["关联卫星"]).strip() for m in matrices}

    # 通道按 (卫星代号, 简称) 索引;通道代号前缀即卫星代号(到第一个'-')
    chan_by_sat_short = {}
    chan_group = {}
    for c in channels:
        code = str(c["通道代号"]).strip()
        satcode = code.split("-")[0]
        short = str(c["通道简称"]).strip() if c["通道简称"] is not None else ""
        chan_by_sat_short.setdefault((satcode, short), []).append(int(c["通道ID"]))
        chan_group[int(c["通道ID"])] = group_by_code.get(str(c["关联通道组"]).strip())

    # 端口按 (矩阵id, ioType, 序号) 索引(供开关解析端口id)
    port_by_key = {}
    for p in ports:
        mid = matrix_by_code.get(str(p["关联矩阵"]).strip())
        if mid is not None and p["序号"] is not None:
            port_by_key[(mid, str(p["输入/输出"]).strip(), int(p["序号"]))] = int(p["端口ID"])

    # 接收通道频率索引:卫星 → [(start, end, pol, beamCode, channelId)]
    grp_info = {int(g["通道组ID"]): g for g in groups}
    rx_chan_index = {}
    for c in channels:
        gid = chan_group.get(int(c["通道ID"]))
        g = grp_info.get(gid)
        if not g or str(g["收/发"]).strip() != "R":
            continue
        try:
            s0, e0 = float(c["通道起始频率"]), float(c["通道终止频率"])
        except (TypeError, ValueError):
            continue
        satcode = str(g["关联卫星代号"]).strip()
        rx_chan_index.setdefault(satcode, []).append((
            s0, e0,
            str(g["极化"]).strip() if g["极化"] is not None else "",
            str(g["波束（天线）代号"]).strip() if g["波束（天线）代号"] is not None else "",
            int(c["通道ID"]),
        ))

    def resolve_channel(blk):
        """块 → 接收通道:同卫星,上行范围落入通道范围;优先极化+波束都匹配。"""
        if not blk:
            return None
        cands = rx_chan_index.get(blk["sat"], [])
        hits = [(s0, e0, pol, beam, cid) for s0, e0, pol, beam, cid in cands
                if s0 - 0.01 <= blk["ustart"] and blk["uend"] <= e0 + 0.01]
        strict = [h for h in hits if h[2] == blk["upol"] and h[3] == blk["ubeam"]]
        pool = strict or [h for h in hits if h[2] == blk["upol"]] or hits
        if not pool:
            return None
        # 多候选时取最窄通道(最贴合)
        pool.sort(key=lambda h: h[1] - h[0])
        return pool[0][4]

    # ══ 002 资源层 ══
    print("生成资源层种子 ...")
    sat_rows = ["(" + ", ".join([
        intv(s["卫星清单ID"]), sq(s["卫星代号"]), sq(s["卫星代号（非标）"]),
        sq(s["卫星名称"]), sq(s["状态"]), sq(s["轨位"]), d_only(s["发射时间"]),
        sq(s["寿命"]), sq(s["制造商"]), sq(s["卫星平台"]), sq(s["覆盖"]), sq(s["有效载荷"]),
    ]) + ")" for s in sats]

    beacon_rows = ["(" + ", ".join([
        intv(b["信标ID"]), sq(b["关联卫星代号"]),
        intv(sat_by_code.get(str(b["关联卫星代号"]).strip())),
        sq(b["频段"]), sq(b["极化"]), num(b["频点"]),
    ]) + ")" for b in beacons]

    group_rows = ["(" + ", ".join([
        intv(g["通道组ID"]), sq(g["通道组代号"]), sq(g["通道组序号"]),
        sq(g["关联卫星代号"]), intv(sat_by_code.get(str(g["关联卫星代号"]).strip())),
        sq(g["波束（天线）名称"]), sq(g["波束（天线）代号"]),
        sq(g["收/发"]), sq(g["极化"]), sq(g["频率"]), intv(g["通道数"]),
        sq(g["主份接收机代码"]), sq(g["一备接收机代码"]), sq(g["二备接收机代码"]),
        sq(g["接收机主备状态"]),
    ]) + ")" for g in groups]

    chan_rows = ["(" + ", ".join([
        intv(c["通道ID"]), sq(c["通道代号"]), sq(c["通道全称"]), sq(c["通道简称"]),
        sq(c["常用名"]), sq(c["关联通道组"]),
        intv(group_by_code.get(str(c["关联通道组"]).strip())),
        num(c["通道带宽"]), num(c["通道起始频率"]), num(c["通道终止频率"]),
    ]) + ")" for c in channels]

    matrix_rows = ["(" + ", ".join([
        intv(m["矩阵ID"]), sq(m["矩阵代码"]), sq(m["关联卫星"]),
        intv(sat_by_code.get(str(m["关联卫星"]).strip())),
        intv(m["类型"]), intv(m["序号"]), intv(m["输入端口数"]), intv(m["输出端口数"]),
        intv(m["生效状态"]), sq(m["备注"]), dt(m["变更时间"]),
    ]) + ")" for m in matrices]

    n_port_amb = 0
    port_rows = []
    for p in ports:
        mid = matrix_by_code.get(str(p["关联矩阵"]).strip())
        satcode = matrix_sat.get(mid, "")
        short = str(p["关联通道代码"]).strip() if p["关联通道代码"] is not None else ""
        cids = chan_by_sat_short.get((satcode, short), [])
        cid = cids[0] if len(cids) == 1 else None  # 多义留 NULL
        if len(cids) > 1:
            n_port_amb += 1
        port_rows.append("(" + ", ".join([
            intv(p["端口ID"]), sq(p["端口代码"]), sq(p["关联矩阵"]), intv(mid),
            sq(p["输入/输出"]), intv(p["序号"]), sq(p["关联通道代码"]), intv(cid),
        ]) + ")")

    switch_rows = []
    for s in switches:
        mid = matrix_by_code.get(str(s["关联矩阵"]).strip())
        in_seq = int(s["入端口号"]) if s["入端口号"] is not None else None
        out_seq = int(s["出端口号"]) if s["出端口号"] is not None else None
        in_pid = port_by_key.get((mid, "I", in_seq)) if in_seq is not None else None
        out_pid = port_by_key.get((mid, "O", out_seq)) if out_seq is not None else None
        switch_rows.append("(" + ", ".join([
            intv(s["开关ID"]), sq(s["开关代码"]), sq(s["关联矩阵"]), intv(mid),
            intv(in_seq), intv(out_seq), intv(in_pid), intv(out_pid),
            sq(s["开关是否可切"]), intv(s["开关状态"]),
            sq(s["主份放大器代码"]), sq(s["一备放大器代码"]), sq(s["二备放大器代码"]),
            sq(s["放大器主备状态"]), dt(s["变更时间"]),
        ]) + ")")

    body2 = "\n".join([
        "-- 002: 资源层种子(卫星/信标/通道组/通道/矩阵/端口/开关)",
        emit("satellite_info",
             ["id", "satelliteCode", "satelliteCodeNonStd", "satelliteName", "statusText",
              "orbitPosition", "launchDate", "designLife", "manufacturer", "platform",
              "coverage", "payload"], sat_rows),
        emit("beacon_info",
             ["id", "satelliteCode", "satelliteId", "band", "polarization", "frequency"],
             beacon_rows),
        emit("channel_group_info",
             ["id", "channelGroupCode", "groupSeq", "satelliteCode", "satelliteId",
              "antennaName", "antennaCode", "txRxType", "polarization", "band",
              "channelCount", "primaryReceiverCode", "backupReceiverCode1",
              "backupReceiverCode2", "receiverActiveStatus"], group_rows),
        emit("channel_info",
             ["id", "channelCode", "channelFullName", "channelShortName", "commonName",
              "channelGroupCode", "channelGroupId", "channelBandwidth",
              "channelStartFreq", "channelEndFreq"], chan_rows),
        emit("switch_matrix_info",
             ["id", "matrixCode", "satelliteCode", "satelliteId", "matrixType", "matrixSeq",
              "inputPortCount", "outputPortCount", "effectiveStatus", "remark", "updateTime"],
             matrix_rows),
        emit("matrix_port_info",
             ["id", "portCode", "matrixCode", "matrixId", "ioType", "portSeq",
              "channelShortName", "channelId"], port_rows),
        emit("matrix_switch_status",
             ["id", "switchCode", "matrixCode", "matrixId", "inputPortSeq", "outputPortSeq",
              "inputPortId", "outputPortId", "switchType", "switchStatus", "primaryAmpCode",
              "backupAmpCode1", "backupAmpCode2", "ampActiveStatus", "updateTime"],
             switch_rows),
    ])

    # ══ 003 状态层 ══
    print("生成状态层种子 ...")

    def block_cols(blk):
        if not blk:
            return ["NULL"] * 11
        return [
            sq(blk["sat"]), intv(sat_by_code.get(blk["sat"])), num(blk["bw"]),
            sq(blk["upol"]), sq(blk["ubeam"]), num(blk["ustart"]), num(blk["uend"]),
            sq(blk["dpol"]), sq(blk["dbeam"]), num(blk["dstart"]), num(blk["dend"]),
        ]

    plan_parsed = {}   # id → blk
    plan_rows = []
    n_plan_unparsed = 0
    for p in plans:
        code = str(p["块代码"]).strip()
        blk = parse_block(code)
        if blk is None:
            n_plan_unparsed += 1
        plan_parsed[int(p["规划记录ID"])] = blk
        cid = resolve_channel(blk)
        plan_rows.append("(" + ", ".join(
            [intv(p["规划记录ID"]), sq(code), sq(p["用途"]), yn(p["记录是否有效"]),
             dt(p["最后更新时间"])] + block_cols(blk) + [intv(cid)]
        ) + ")")

    # 规划块按卫星索引,供分配块归属解析
    plan_index = {}
    for pid, blk in plan_parsed.items():
        if blk:
            plan_index.setdefault(blk["sat"], []).append((pid, blk))

    def resolve_planning(blk):
        """分配块 → 规划块:同星同上行极化同波束,范围包含;多候选取最窄。"""
        if not blk:
            return None
        cands = [(pid, q) for pid, q in plan_index.get(blk["sat"], [])
                 if q["upol"] == blk["upol"] and q["ubeam"] == blk["ubeam"]
                 and q["ustart"] - 0.01 <= blk["ustart"] and blk["uend"] <= q["uend"] + 0.01]
        if not cands:
            return None
        cands.sort(key=lambda t: t[1]["uend"] - t[1]["ustart"])
        return cands[0][0]

    alloc_by_code = {}  # trim 后块代码 → 占用记录ID(首个)
    alloc_rows = []
    n_alloc_orphan = 0
    for a in allocs:
        code = str(a["块代码"]).strip()
        blk = parse_block(code)
        pid = resolve_planning(blk)
        if pid is None:
            n_alloc_orphan += 1
        cid = resolve_channel(blk)
        aid = int(a["占用记录ID"])
        alloc_by_code.setdefault(code, aid)
        alloc_rows.append("(" + ", ".join(
            [intv(aid), sq(code), yn(a["分配是否有效"]), dt(a["最后更新时间"])]
            + block_cols(blk) + [intv(pid), intv(cid)]
        ) + ")")

    body3 = "\n".join([
        "-- 003: 状态层种子(通道规划状态/通道分配状态)",
        emit("channel_planning_status",
             ["id", "blockCode", "usageType", "isValid", "updateTime",
              "satelliteCode", "satelliteId", "bandwidth",
              "uplinkPolarization", "uplinkBeam", "uplinkStartFreq", "uplinkEndFreq",
              "downlinkPolarization", "downlinkBeam", "downlinkStartFreq", "downlinkEndFreq",
              "channelId"], plan_rows),
        emit("channel_allocation_status",
             ["id", "blockCode", "isValid", "updateTime",
              "satelliteCode", "satelliteId", "bandwidth",
              "uplinkPolarization", "uplinkBeam", "uplinkStartFreq", "uplinkEndFreq",
              "downlinkPolarization", "downlinkBeam", "downlinkStartFreq", "downlinkEndFreq",
              "planningBlockId", "channelId"], alloc_rows),
    ])

    # ══ 004 业务层 ══
    print("生成业务层种子 ...")
    cust_rows = ["(" + ", ".join([
        sq(k["客户ID"]), sq(k["客户全称"]), sq(k["统一社会信用代码"]),
        intv(k["客户状态"]), dt(k["建档时间"]), dt(k["最后更新时间"]),
    ]) + ")" for k in customers]

    user_rows = ["(" + ", ".join([
        intv(u["用户ID"]), sq(u["所属客户ID"]), sq(u["所属客户全称"]),
        intv(u["用户状态"]), dt(u["建档时间"]), dt(u["最后更新时间"]),
    ]) + ")" for u in users]

    user_customer = {int(u["用户ID"]): str(u["所属客户ID"]).strip()
                     for u in users if u["所属客户ID"] is not None}

    contract_rows = []
    for c in contracts:
        uid = int(c["用户号"]) if c["用户号"] is not None else None
        contract_rows.append("(" + ", ".join([
            intv(c["合约ID"]), sq(c["所属客户"]), sq(user_customer.get(uid)),
            intv(uid), sq(c["所属主订单"]), sq(c["签约带宽商品"]), sq(c["商品类型"]),
            num(c["签约带宽权益"]), intv(c["可分频率块数量"]), num(c["期数"]),
            num(c["签约金额"]), dt(c["开通时间"]), dt(c["到期时间"]), dt(c["最后更新时间"]),
        ]) + ")")

    n_dlv_miss = 0
    delivery_rows = []
    for d in deliveries:
        code = str(d["频率块代码"]).strip() if d["频率块代码"] is not None else None
        aid = alloc_by_code.get(code)
        if code and aid is None:
            n_dlv_miss += 1
        delivery_rows.append("(" + ", ".join([
            intv(d["交付记录ID"]), intv(d["关联合约ID"]), sq(code), intv(aid),
            sq(d["独占/共享"]), sq(d["频率块所属卫星"]),
            intv(sat_by_code.get(str(d["频率块所属卫星"]).strip())
                 if d["频率块所属卫星"] is not None else None),
            num(d["交付频率块带宽"]), sq(d["动作"]), dt(d["时间"]),
            sq(d["受理人员"]), sq(d["登记人员"]),
        ]) + ")")

    n_usage_miss = 0
    usage_rows = []
    for u in usages:
        code = str(u["频率块代码"]).strip() if u["频率块代码"] is not None else None
        aid = alloc_by_code.get(code)
        if code and aid is None:
            n_usage_miss += 1
        usage_rows.append("(" + ", ".join([
            intv(u["使用记录ID"]), intv(u["关联载波ID"]), sq(code), intv(aid),
            sq(u["独占/共享"]), sq(u["频率块所属卫星"]),
            intv(sat_by_code.get(str(u["频率块所属卫星"]).strip())
                 if u["频率块所属卫星"] is not None else None),
            num(u["频率块带宽"]), sq(u["动作"]), dt(u["时间"]),
            sq(u["受理人员"]), sq(u["登记人员"]),
        ]) + ")")

    body4 = "\n".join([
        "-- 004: 业务层种子(客户/用户/合约/交付记录/自有载波使用记录)",
        emit("customer_info",
             ["customerCode", "customerName", "creditCode", "status",
              "createdTime", "updateTime"], cust_rows),
        emit("user_info",
             ["id", "customerCode", "customerName", "status", "createdTime", "updateTime"],
             user_rows),
        emit("bandwidth_contract_info",
             ["id", "customerName", "customerCode", "userId", "mainOrderCode",
              "productName", "productType", "bandwidthMHz", "divisibleBlockCount",
              "periods", "amount", "startTime", "endTime", "updateTime"], contract_rows),
        emit("contract_delivery_record",
             ["id", "contractId", "blockCode", "allocationId", "exclusiveType",
              "satelliteCode", "satelliteId", "bandwidth", "action", "actionTime",
              "handler", "registrar"], delivery_rows),
        emit("own_carrier_usage_record",
             ["id", "carrierId", "blockCode", "allocationId", "exclusiveType",
              "satelliteCode", "satelliteId", "bandwidth", "action", "actionTime",
              "handler", "registrar"], usage_rows),
    ])

    # ── 写出 ──
    write("001_create_schema_v6.sql", SCHEMA)
    write("002_seed_resource_layer.sql", body2)
    write("003_seed_status_layer.sql", body3)
    write("004_seed_business_layer.sql", body4)

    print("\n解析统计:")
    print(f"  端口简称多义(channelId=NULL): {n_port_amb}")
    print(f"  规划块代码无法解析: {n_plan_unparsed}")
    print(f"  分配块未落入任何规划块(planningBlockId=NULL): {n_alloc_orphan}")
    print(f"  交付记录块代码未命中分配状态: {n_dlv_miss}")
    print(f"  自有使用记录块代码未命中分配状态: {n_usage_miss}")
    print("完成。")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
