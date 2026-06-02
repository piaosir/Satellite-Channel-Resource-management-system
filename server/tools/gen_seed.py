"""
种子迁移生成器
==============
读取 Excel（通道基础信息和清单），生成 server/migrations/002…007.sql。
所有迁移为静态文件并提交到仓库；本脚本仅供复现。

用法：
    python server/tools/gen_seed.py                 # 用 server/tools/channel_data.xlsx
    python server/tools/gen_seed.py --excel <路径>  # 指定 Excel

关键约束（针对 migrate.py 按 ';' 切分的实现）：
  - 字符串中的 ASCII ';' 一律替换为全角 '；'，避免误切分语句
  - 单引号转义为 ''，'#N/A'/空白 → NULL
  - 多行 INSERT 每批 <= 200 行
  - 外键 id（satelliteId/switchId/frequencyBlockId）在迁移末尾用 UPDATE…JOIN 按自然代号解析，
    生成期不连库
"""
import argparse
import datetime as dt
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
MIGRATIONS = os.path.join(HERE, "..", "migrations")
DEFAULT_XLSX = os.path.join(HERE, "channel_data.xlsx")

SHEET = {
    "sat":      "卫星表",
    "twt":      "TWT实时状态（新）",
    "chattr":   "通道属性信息表(新)",
    "swgroup":  "开关组信息表（新）",
    "product":  "商品实例清单",
    "contract": "合约记录 (新-20260601-1647)",
}

BATCH = 200


# ── 字面量工具 ──────────────────────────────────────────────────
def _blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() in ("", "#N/A", "#n/a"))


def txt(v) -> str:
    if _blank(v):
        return "NULL"
    s = str(v).strip()
    s = s.replace(";", "；").replace("'", "''")
    return "'" + s + "'"


def num(v) -> str:
    """数值字面量；非数值 → NULL。"""
    if _blank(v):
        return "NULL"
    if isinstance(v, bool):
        return "NULL"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return repr(v)
    s = str(v).strip().replace(",", "")
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else repr(f)
    except ValueError:
        return "NULL"


def intval(v) -> str:
    if _blank(v):
        return "NULL"
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return "NULL"


def dtime(v) -> str:
    if _blank(v):
        return "NULL"
    if isinstance(v, dt.datetime):
        return "'" + v.strftime("%Y-%m-%d %H:%M:%S") + "'"
    if isinstance(v, dt.date):
        return "'" + v.strftime("%Y-%m-%d 00:00:00") + "'"
    return txt(v)


def date_only(v) -> str:
    if _blank(v):
        return "NULL"
    if isinstance(v, (dt.datetime, dt.date)):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    return txt(v)


# ── 通用工具 ────────────────────────────────────────────────────
def header_index(ws):
    """返回 {表头文本: 列下标}，取首行。"""
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h).strip(): i for i, h in enumerate(first) if h is not None}


def data_rows(ws):
    return list(ws.iter_rows(min_row=2, values_only=True))


def cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def emit_inserts(table, cols, rows_sql):
    """把若干 (val1,val2,...) 字符串行拼成多条多行 INSERT。"""
    out = []
    col_list = "(" + ", ".join(f"`{c}`" for c in cols) + ")"
    for i in range(0, len(rows_sql), BATCH):
        chunk = rows_sql[i:i + BATCH]
        out.append(f"INSERT INTO `{table}` {col_list} VALUES\n  " +
                   ",\n  ".join(chunk) + ";")
    return "\n".join(out)


def write_migration(name, body):
    path = os.path.join(MIGRATIONS, name)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(body.rstrip() + "\n")
    print(f"  写出 {name}  ({len(body.splitlines())} 行)")


# ════════════════════════════════════════════════════════════════
def gen_satellite(wb):
    ws = wb[SHEET["sat"]]
    h = header_index(ws)
    norm = lambda c: str(c).replace("-", "")
    spec = [  # (db_col, excel_header, fn)
        ("orbitPosition", "轨道位置", txt),
        ("statusText", "状态", txt),
        ("coverage", "覆盖范围", txt),
        ("transponderCount", "转发器数量", txt),
        ("beacon", "信标（MHz）", txt),
        ("polarization", "极化方式", txt),
        ("launchDate", "发射时间", date_only),
        ("designLife", "设计寿命", txt),
        ("ownership", "卫星归属", txt),
        ("manufacturer", "制造商", txt),
        ("platform", "卫星平台", txt),
        ("attitudeStabilization", "姿态稳定", txt),
        ("stationKeepingAccuracy", "位保精度", txt),
        ("remark", "备注", txt),
    ]
    lines = [
        "-- 002: 扩展 satellite_info，补充卫星档案字段（MySQL 8，无 IF NOT EXISTS）",
        "ALTER TABLE `satellite_info`",
        "  ADD COLUMN `orbitPosition` VARCHAR(64) NULL COMMENT '轨道位置',",
        "  ADD COLUMN `statusText` VARCHAR(32) NULL COMMENT '状态：在轨运营/停止服务/离轨/在建',",
        "  ADD COLUMN `coverage` TEXT NULL COMMENT '覆盖范围',",
        "  ADD COLUMN `transponderCount` VARCHAR(255) NULL COMMENT '转发器数量',",
        "  ADD COLUMN `beacon` TEXT NULL COMMENT '信标(MHz)',",
        "  ADD COLUMN `polarization` VARCHAR(64) NULL COMMENT '极化方式',",
        "  ADD COLUMN `launchDate` DATE NULL COMMENT '发射时间',",
        "  ADD COLUMN `designLife` VARCHAR(32) NULL COMMENT '设计寿命',",
        "  ADD COLUMN `ownership` VARCHAR(32) NULL COMMENT '卫星归属：自有/代理',",
        "  ADD COLUMN `manufacturer` VARCHAR(128) NULL COMMENT '制造商',",
        "  ADD COLUMN `platform` VARCHAR(128) NULL COMMENT '卫星平台',",
        "  ADD COLUMN `attitudeStabilization` VARCHAR(64) NULL COMMENT '姿态稳定',",
        "  ADD COLUMN `stationKeepingAccuracy` VARCHAR(128) NULL COMMENT '位保精度',",
        "  ADD COLUMN `remark` TEXT NULL COMMENT '备注';",
        "",
    ]
    n_upd = n_ins = 0
    for row in data_rows(ws):
        code = cell(row, h.get("卫星代号"))
        name = cell(row, h.get("卫星名称"))
        if _blank(code):
            continue
        nc = norm(code)
        sets = [f"`{db}` = {fn(cell(row, h.get(hd)))}" for db, hd, fn in spec]
        sets.append(f"`satelliteName` = {txt(name)}")
        lines.append(
            f"UPDATE `satellite_info` SET {', '.join(sets)} "
            f"WHERE REPLACE(`satelliteCode`,'-','') = '{nc}';"
        )
        n_upd += 1
        # 库中缺失的代号（如 CS27）才插入
        ins_cols = ["satelliteCode", "satelliteName"] + [db for db, _, _ in spec]
        ins_vals = [txt(code), txt(name)] + [fn(cell(row, h.get(hd))) for _, hd, fn in spec]
        lines.append(
            f"INSERT INTO `satellite_info` (" + ", ".join(f"`{c}`" for c in ins_cols) + ")\n"
            f"  SELECT " + ", ".join(ins_vals) + " FROM DUAL\n"
            f"  WHERE NOT EXISTS (SELECT 1 FROM `satellite_info` "
            f"WHERE REPLACE(`satelliteCode`,'-','') = '{nc}');"
        )
        n_ins += 1
    print(f"  satellite: {n_upd} UPDATE + {n_ins} 条幂等 INSERT")
    write_migration("002_extend_satellite_info.sql", "\n".join(lines))


def gen_twt(wb):
    ws = wb[SHEET["twt"]]
    h = header_index(ws)
    cols = ["twtCodeLong", "twtCodeShort", "satelliteCode", "unitCode",
            "onOff", "mutingStatus", "gainMode", "gainLevel"]
    rows_sql = []
    for row in data_rows(ws):
        if _blank(cell(row, h.get("TWT代码（长）"))):
            continue
        rows_sql.append("(" + ", ".join([
            txt(cell(row, h.get("TWT代码（长）"))),
            txt(cell(row, h.get("TWT代码（短）"))),
            txt(cell(row, h.get("所属卫星"))),
            txt(cell(row, h.get("单机代号"))),
            txt(cell(row, h.get("ON/OFF"))),
            txt(cell(row, h.get("Muting Status"))),
            txt(cell(row, h.get("FGM/ALC"))),
            intval(cell(row, h.get("档位"))),
        ]) + ")")
    ddl = """-- 003: 行波管 TWT 实时状态表 + 种子
CREATE TABLE IF NOT EXISTS `twt_realtime_status` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '主键id',
  `twtCodeLong` VARCHAR(128) NULL COMMENT 'TWT代码(长)',
  `twtCodeShort` VARCHAR(64) NULL COMMENT 'TWT代码(短)',
  `satelliteCode` VARCHAR(64) NULL COMMENT '所属卫星代号',
  `satelliteId` INT NULL COMMENT '所属卫星id(解析)',
  `unitCode` VARCHAR(64) NULL COMMENT '单机代号',
  `onOff` VARCHAR(16) NULL COMMENT 'ON/OFF',
  `mutingStatus` VARCHAR(32) NULL COMMENT 'Muting Status',
  `gainMode` VARCHAR(16) NULL COMMENT 'FGM/ALC',
  `gainLevel` INT NULL COMMENT '档位',
  `statusUpdateTime` BIGINT NULL COMMENT '状态修改时间(毫秒)',
  PRIMARY KEY (`id`),
  INDEX `idx_twt_satellite_id` (`satelliteId`),
  INDEX `idx_twt_satellite_code` (`satelliteCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='行波管TWT实时状态表';
"""
    body = ddl + "\n" + emit_inserts("twt_realtime_status", cols, rows_sql) + "\n\n" + (
        "UPDATE `twt_realtime_status` t "
        "JOIN `satellite_info` s ON REPLACE(s.`satelliteCode`,'-','') = REPLACE(t.`satelliteCode`,'-','') "
        "SET t.`satelliteId` = s.`id`;"
    )
    print(f"  twt: {len(rows_sql)} 行")
    write_migration("003_create_twt_realtime_status.sql", body)


def gen_channel_attr(wb):
    ws = wb[SHEET["chattr"]]
    h = header_index(ws)
    cols = ["switchCode", "matrixCode", "inputPortSeq", "outputPortSeq",
            "inputChannelShortName", "outputChannelShortName", "gainMode",
            "currentLevel", "startLevel", "maxLevel", "levelStep",
            "startSfdRef", "currentSfd", "satelliteCode"]

    def gain_mode(v):
        s = "" if v is None else str(v).strip()
        return txt(s) if s in ("FGM", "ALC") else "NULL"

    rows_sql = []
    for row in data_rows(ws):
        if _blank(cell(row, h.get("开关代码"))):
            continue
        rows_sql.append("(" + ", ".join([
            txt(cell(row, h.get("开关代码"))),
            txt(cell(row, h.get("所属矩阵"))),
            intval(cell(row, h.get("入端口序号"))),
            intval(cell(row, h.get("出端口序号"))),
            txt(cell(row, h.get('索引"入端口通道“'))),
            txt(cell(row, h.get('索引"出端口通道"'))),
            gain_mode(cell(row, h.get("FGM/ALC"))),
            intval(cell(row, h.get("当前档位"))),
            intval(cell(row, h.get("起始档位"))),
            intval(cell(row, h.get("最高档位"))),
            txt(cell(row, h.get("档位步进（dB/step）"))),
            txt(cell(row, h.get("起始档位SFD(ref)"))),
            txt(cell(row, h.get("当前档位SFD（dBW/m^2）"))),
            txt(cell(row, h.get("所属卫星"))),
        ]) + ")")
    ddl = """-- 004: 通道属性(增益/SFD)表 + 种子
CREATE TABLE IF NOT EXISTS `channel_attribute_info` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '主键id',
  `switchCode` VARCHAR(128) NULL COMMENT '开关代码',
  `matrixCode` VARCHAR(128) NULL COMMENT '所属矩阵',
  `inputPortSeq` INT NULL COMMENT '入端口序号',
  `outputPortSeq` INT NULL COMMENT '出端口序号',
  `inputChannelShortName` VARCHAR(64) NULL COMMENT '入端口通道',
  `outputChannelShortName` VARCHAR(64) NULL COMMENT '出端口通道',
  `gainMode` VARCHAR(16) NULL COMMENT 'FGM/ALC',
  `currentLevel` INT NULL COMMENT '当前档位',
  `startLevel` INT NULL COMMENT '起始档位',
  `maxLevel` INT NULL COMMENT '最高档位',
  `levelStep` VARCHAR(16) NULL COMMENT '档位步进(dB/step)',
  `startSfdRef` VARCHAR(32) NULL COMMENT '起始档位SFD(ref)',
  `currentSfd` VARCHAR(64) NULL COMMENT '当前档位SFD(dBW/m^2)',
  `satelliteCode` VARCHAR(64) NULL COMMENT '所属卫星代号',
  `satelliteId` INT NULL COMMENT '所属卫星id(解析)',
  `switchId` INT NULL COMMENT '开关id(解析)',
  PRIMARY KEY (`id`),
  INDEX `idx_chattr_satellite_id` (`satelliteId`),
  INDEX `idx_chattr_switch_code` (`switchCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='通道属性(增益/SFD)表';
"""
    body = ddl + "\n" + emit_inserts("channel_attribute_info", cols, rows_sql) + "\n\n" + (
        "UPDATE `channel_attribute_info` a "
        "JOIN `satellite_info` s ON REPLACE(s.`satelliteCode`,'-','') = REPLACE(a.`satelliteCode`,'-','') "
        "SET a.`satelliteId` = s.`id`;\n"
        "UPDATE `channel_attribute_info` a "
        "JOIN `matrix_switch_status` w ON w.`switchCode` = a.`switchCode` "
        "SET a.`switchId` = w.`id`;"
    )
    print(f"  channel_attr: {len(rows_sql)} 行")
    write_migration("004_create_channel_attribute_info.sql", body)


def _sat_codes(wb):
    ws = wb[SHEET["sat"]]
    h = header_index(ws)
    codes = []
    for row in data_rows(ws):
        c = cell(row, h.get("卫星代号"))
        if not _blank(c):
            codes.append(str(c).strip())
    # 长前缀优先，且同时考虑去横杠形式
    return sorted(set(codes), key=len, reverse=True)


def _derive_sat(matrix_code, sat_codes):
    if _blank(matrix_code):
        return None
    mc = str(matrix_code).strip()
    for c in sat_codes:
        if mc == c or mc.startswith(c + "-"):
            return c
        nc = c.replace("-", "")
        if mc == nc or mc.startswith(nc + "-"):
            return c
    return None


def gen_switch_group(wb):
    ws = wb[SHEET["swgroup"]]
    h = header_index(ws)
    sat_codes = _sat_codes(wb)
    cols = ["switchGroupCode", "switchCode", "matrixCode", "inputPortSeq",
            "outputPortSeq", "inputChannelShortName", "outputChannelShortName",
            "switchStatus", "switchType", "checkRule", "satelliteCode"]
    check_hdr = next((k for k in h if k.startswith("校验")), None)
    rows_sql = []
    for row in data_rows(ws):
        if _blank(cell(row, h.get("开关代码"))):
            continue
        mc = cell(row, h.get("所属矩阵"))
        rows_sql.append("(" + ", ".join([
            txt(cell(row, h.get("开关组代码"))),
            txt(cell(row, h.get("开关代码"))),
            txt(mc),
            intval(cell(row, h.get("入端口序号"))),
            intval(cell(row, h.get("出端口序号"))),
            txt(cell(row, h.get('索引"入端口通道“'))),
            txt(cell(row, h.get('索引"出端口通道"'))),
            intval(cell(row, h.get("开关状态 1-通，0-断"))),
            txt(cell(row, h.get("常通/可切"))),
            txt(cell(row, h.get(check_hdr)) if check_hdr else None),
            txt(_derive_sat(mc, sat_codes)),
        ]) + ")")
    ddl = """-- 005: 开关组信息表 + 种子
CREATE TABLE IF NOT EXISTS `switch_group_info` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '主键id',
  `switchGroupCode` VARCHAR(128) NULL COMMENT '开关组代码',
  `switchCode` VARCHAR(128) NULL COMMENT '开关代码',
  `matrixCode` VARCHAR(128) NULL COMMENT '所属矩阵',
  `inputPortSeq` INT NULL COMMENT '入端口序号',
  `outputPortSeq` INT NULL COMMENT '出端口序号',
  `inputChannelShortName` VARCHAR(64) NULL COMMENT '入端口通道',
  `outputChannelShortName` VARCHAR(64) NULL COMMENT '出端口通道',
  `switchStatus` INT NULL COMMENT '开关状态 1-通 0-断',
  `switchType` VARCHAR(32) NULL COMMENT '常通/可切',
  `checkRule` VARCHAR(255) NULL COMMENT '校验：模拟开关组仅允许一个置1，数字组无限制',
  `satelliteCode` VARCHAR(64) NULL COMMENT '所属卫星代号(由矩阵推导)',
  `satelliteId` INT NULL COMMENT '所属卫星id(解析)',
  PRIMARY KEY (`id`),
  INDEX `idx_swgroup_satellite_id` (`satelliteId`),
  INDEX `idx_swgroup_code` (`switchGroupCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='开关组信息表';
"""
    body = ddl + "\n" + emit_inserts("switch_group_info", cols, rows_sql) + "\n\n" + (
        "UPDATE `switch_group_info` g "
        "JOIN `satellite_info` s ON REPLACE(s.`satelliteCode`,'-','') = REPLACE(g.`satelliteCode`,'-','') "
        "SET g.`satelliteId` = s.`id`;"
    )
    print(f"  switch_group: {len(rows_sql)} 行")
    write_migration("005_create_switch_group_info.sql", body)


def gen_product(wb):
    ws = wb[SHEET["product"]]
    h = header_index(ws)
    # (db_col, excel_header, fn)
    spec = [
        ("productInstanceCode", "商品实例编号", txt),
        ("subOrderCode", "子订单编号", txt),
        ("productName", "商品名称", txt),
        ("instanceType", "商品实例类型", txt),
        ("unitPrice", "商品实例成交单价", num),
        ("contractPeriod", "订购期数/合约期", txt),
        ("planStartTime", "计划开始时间", dtime),
        ("planEndTime", "计划结束时间", dtime),
        ("fulfillStatus", "商品实例履约状态", txt),
        ("subOrderCategory", "子订单类别", txt),
        ("mainOrderCode", "关联主订单编号", txt),
        ("contractNo", "合同号", txt),
        ("partyA", "甲方信息", txt),
        ("groupName", "分组情况", txt),
        ("sales", "销售", txt),
        ("reporter", "填报人", txt),
        ("subOrderAmount", "子订单金额", num),
        ("mainOrderAmount", "主订单金额", num),
        ("bandwidthMHz", "带宽(MHz)", num),
        ("satelliteCode", "卫星", txt),
        ("frequencyBlockCode2", "块编号", txt),
        ("exclusiveType", "独占/共享", txt),
        ("remark", "备注", txt),
    ]
    cols = [c for c, _, _ in spec]
    rows_sql = []
    for row in data_rows(ws):
        if _blank(cell(row, h.get("商品实例编号"))):
            continue
        rows_sql.append("(" + ", ".join(fn(cell(row, h.get(hd))) for _, hd, fn in spec) + ")")
    ddl = """-- 006: 商品实例清单表 + 种子
CREATE TABLE IF NOT EXISTS `product_instance` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '主键id',
  `productInstanceCode` VARCHAR(128) NULL COMMENT '商品实例编号',
  `subOrderCode` VARCHAR(128) NULL COMMENT '子订单编号',
  `productName` VARCHAR(255) NULL COMMENT '商品名称',
  `instanceType` VARCHAR(32) NULL COMMENT '商品实例类型',
  `unitPrice` DECIMAL(18,2) NULL COMMENT '成交单价',
  `contractPeriod` VARCHAR(32) NULL COMMENT '订购期数/合约期',
  `planStartTime` DATETIME NULL COMMENT '计划开始时间',
  `planEndTime` DATETIME NULL COMMENT '计划结束时间',
  `fulfillStatus` VARCHAR(32) NULL COMMENT '履约状态',
  `subOrderCategory` VARCHAR(32) NULL COMMENT '子订单类别',
  `mainOrderCode` VARCHAR(128) NULL COMMENT '关联主订单编号',
  `contractNo` VARCHAR(128) NULL COMMENT '合同号',
  `partyA` VARCHAR(255) NULL COMMENT '甲方信息',
  `groupName` VARCHAR(64) NULL COMMENT '分组情况',
  `sales` VARCHAR(64) NULL COMMENT '销售',
  `reporter` VARCHAR(64) NULL COMMENT '填报人',
  `subOrderAmount` DECIMAL(18,2) NULL COMMENT '子订单金额',
  `mainOrderAmount` DECIMAL(18,2) NULL COMMENT '主订单金额',
  `bandwidthMHz` DECIMAL(12,2) NULL COMMENT '带宽(MHz)',
  `satelliteCode` VARCHAR(64) NULL COMMENT '卫星',
  `frequencyBlockCode2` VARCHAR(255) NULL COMMENT '频率块代码-2(块编号)',
  `exclusiveType` VARCHAR(32) NULL COMMENT '独占/共享',
  `remark` TEXT NULL COMMENT '备注',
  PRIMARY KEY (`id`),
  INDEX `idx_pi_code` (`productInstanceCode`),
  INDEX `idx_pi_partya` (`partyA`),
  INDEX `idx_pi_satellite_code` (`satelliteCode`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='商品实例清单表';
"""
    body = ddl + "\n" + emit_inserts("product_instance", cols, rows_sql)
    print(f"  product_instance: {len(rows_sql)} 行")
    write_migration("006_create_product_instance.sql", body)


def gen_contract(wb):
    ws = wb[SHEET["contract"]]
    h = header_index(ws)
    spec = [
        ("remarkInfo", "备注信息", txt),
        ("productInstanceId", "商品实例id", txt),
        ("subOrderCode", "所属子订单", txt),
        ("partyA", "甲方", txt),
        ("productName", "商品名称", txt),
        ("contractNo", "合同编号", txt),
        ("remark", "备注", txt),
        ("frequencyBlockCode2", "频率块代码", txt),
        ("exclusiveType", "独占/共享关系", txt),
        ("usedBandwidth", "使用带宽", num),
        ("startTime", "开始时间", dtime),
        ("endTime", "结束时间", dtime),
        ("satelliteCode", "卫星", txt),
        ("uplinkBeamCode", "上行波束代码", txt),
        ("uplinkPolarization", "上行极化", txt),
        ("uplinkStartFreq", "上行起频率", num),
        ("uplinkEndFreq", "上行止频率", num),
        ("downlinkBeamCode", "下行波束代码", txt),
        ("downlinkPolarization", "下行极化", txt),
        ("downlinkStartFreq", "下行起频率", num),
        ("downlinkEndFreq", "下行止频率", num),
    ]
    cols = [c for c, _, _ in spec]
    rows_sql = []
    for row in data_rows(ws):
        if _blank(cell(row, h.get("商品实例id"))) and _blank(cell(row, h.get("频率块代码"))):
            continue
        rows_sql.append("(" + ", ".join(fn(cell(row, h.get(hd))) for _, hd, fn in spec) + ")")
    ddl = """-- 007: 合约记录表(新) + 种子
CREATE TABLE IF NOT EXISTS `contract_record` (
  `id` INT NOT NULL AUTO_INCREMENT COMMENT '主键id',
  `remarkInfo` VARCHAR(255) NULL COMMENT '备注信息',
  `productInstanceId` VARCHAR(128) NULL COMMENT '商品实例id',
  `subOrderCode` VARCHAR(128) NULL COMMENT '所属子订单',
  `partyA` VARCHAR(255) NULL COMMENT '甲方',
  `productName` VARCHAR(255) NULL COMMENT '商品名称',
  `contractNo` VARCHAR(128) NULL COMMENT '合同编号',
  `remark` TEXT NULL COMMENT '备注',
  `frequencyBlockCode2` VARCHAR(255) NULL COMMENT '频率块代码',
  `exclusiveType` VARCHAR(32) NULL COMMENT '独占/共享关系',
  `usedBandwidth` DECIMAL(12,2) NULL COMMENT '使用带宽',
  `startTime` DATETIME NULL COMMENT '开始时间',
  `endTime` DATETIME NULL COMMENT '结束时间',
  `satelliteCode` VARCHAR(64) NULL COMMENT '卫星',
  `uplinkBeamCode` VARCHAR(64) NULL COMMENT '上行波束代码',
  `uplinkPolarization` VARCHAR(16) NULL COMMENT '上行极化',
  `uplinkStartFreq` DECIMAL(12,2) NULL COMMENT '上行起频率',
  `uplinkEndFreq` DECIMAL(12,2) NULL COMMENT '上行止频率',
  `downlinkBeamCode` VARCHAR(64) NULL COMMENT '下行波束代码',
  `downlinkPolarization` VARCHAR(16) NULL COMMENT '下行极化',
  `downlinkStartFreq` DECIMAL(12,2) NULL COMMENT '下行起频率',
  `downlinkEndFreq` DECIMAL(12,2) NULL COMMENT '下行止频率',
  `satelliteId` INT NULL COMMENT '卫星id(解析)',
  `frequencyBlockId` INT NULL COMMENT '频率块id(解析)',
  PRIMARY KEY (`id`),
  INDEX `idx_contract_partya` (`partyA`),
  INDEX `idx_contract_satellite_id` (`satelliteId`),
  INDEX `idx_contract_fbcode2` (`frequencyBlockCode2`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='合约记录表(新)';
"""
    body = ddl + "\n" + emit_inserts("contract_record", cols, rows_sql) + "\n\n" + (
        "UPDATE `contract_record` c "
        "JOIN `satellite_info` s ON REPLACE(s.`satelliteCode`,'-','') = REPLACE(c.`satelliteCode`,'-','') "
        "SET c.`satelliteId` = s.`id`;\n"
        "UPDATE `contract_record` c "
        "JOIN `frequency_block_realtime_status` f ON f.`frequencyBlockCode2` = c.`frequencyBlockCode2` "
        "SET c.`frequencyBlockId` = f.`id`;"
    )
    print(f"  contract_record: {len(rows_sql)} 行")
    write_migration("007_create_contract_record.sql", body)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default=DEFAULT_XLSX)
    args = ap.parse_args()
    print(f"读取 Excel：{args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    gen_satellite(wb)
    gen_twt(wb)
    gen_channel_attr(wb)
    gen_switch_group(wb)
    gen_product(wb)
    gen_contract(wb)
    print("完成。")


if __name__ == "__main__":
    main()
